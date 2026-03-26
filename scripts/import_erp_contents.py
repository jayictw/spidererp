import argparse
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_HEADERS = [
    "操作",
    "型号",
    "倉庫",
    "品牌",
    "批次",
    "庫存量",
    "積压天数",
    "币种",
    "usd進貨价",
    "含稅進貨价",
    "產地",
    "人民币未稅貨值",
    "人民币含稅貨值",
    "稅点",
    "最低售价",
    "含稅售价",
    "標準售价",
    "標準包裝",
    "最小包裝",
    "入庫时間",
    "包裝類型",
    "生命周期",
    "商品分類",
    "供應狀態",
]


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Import ERP內容.xlsx into sol.db for ERP/history reader.")
    p.add_argument("--input", default=r"C:/Users/PC/Desktop/ERP內容.xlsx", help="ERP workbook path.")
    p.add_argument("--db-path", default=r"F:/Jay_ic_tw/sol.db", help="SQLite database path.")
    return p.parse_args()


def connect_db(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS erp_inventory (
          erp_inventory_id INTEGER PRIMARY KEY AUTOINCREMENT,
          import_batch_id TEXT NOT NULL,
          row_no INTEGER,
          action_type TEXT,
          model TEXT NOT NULL,
          warehouse TEXT,
          brand TEXT,
          batch_no TEXT,
          stock_qty REAL,
          aging_days REAL,
          currency TEXT,
          usd_purchase_price REAL,
          taxed_purchase_price REAL,
          origin TEXT,
          inventory_value_cny_ex_tax REAL,
          inventory_value_cny_tax REAL,
          tax_rate REAL,
          floor_sale_price REAL,
          taxed_sale_price REAL,
          standard_sale_price REAL,
          standard_pack_qty REAL,
          min_pack_qty REAL,
          inbound_date TEXT,
          package_type TEXT,
          lifecycle TEXT,
          product_category TEXT,
          supply_status TEXT,
          source_file TEXT,
          imported_at TEXT NOT NULL
        )
        """
    )
    conn.commit()


def safe_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None


def normalize_date(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value).strip()


def detect_workbook_path(raw_path: str) -> Path:
    p = Path(raw_path)
    if p.exists():
        return p
    desktop = Path(r"C:/Users/PC/Desktop")
    matches = [child for child in desktop.iterdir() if child.suffix.lower() == ".xlsx" and child.name.startswith("ERP")]
    if not matches:
        raise FileNotFoundError(f"ERP workbook not found: {raw_path}")
    return matches[0]


def main() -> int:
    args = parse_args()
    input_path = detect_workbook_path(args.input)
    db_path = Path(args.db_path)

    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        print("[error] workbook has insufficient rows")
        return 2

    headers = [str(x).strip() if x is not None else "" for x in rows[0][1:25]]
    print(f"[input] file={input_path}")
    print(f"[input] sheet={wb.sheetnames[0]}")
    print(f"[input] headers={headers}")

    batch_id = datetime.now().strftime("erp_import_%Y%m%d_%H%M%S")
    imported_at = datetime.now().isoformat(timespec="seconds")

    inserted = 0
    with connect_db(db_path) as conn:
        ensure_table(conn)
        conn.execute("DELETE FROM erp_inventory WHERE source_file=?", (str(input_path),))
        for row in rows[2:]:
            if not row or len(row) < 25:
                continue
            model = str(row[2]).strip() if row[2] is not None else ""
            if not model:
                continue
            conn.execute(
                """
                INSERT INTO erp_inventory(
                  import_batch_id, row_no, action_type, model, warehouse, brand, batch_no,
                  stock_qty, aging_days, currency, usd_purchase_price, taxed_purchase_price,
                  origin, inventory_value_cny_ex_tax, inventory_value_cny_tax, tax_rate,
                  floor_sale_price, taxed_sale_price, standard_sale_price, standard_pack_qty,
                  min_pack_qty, inbound_date, package_type, lifecycle, product_category,
                  supply_status, source_file, imported_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    batch_id,
                    row[0],
                    row[1],
                    model.upper(),
                    row[3],
                    row[4],
                    str(row[5]).strip() if row[5] is not None else "",
                    safe_float(row[6]),
                    safe_float(row[7]),
                    row[8],
                    safe_float(row[9]),
                    safe_float(row[10]),
                    row[11],
                    safe_float(row[12]),
                    safe_float(row[13]),
                    safe_float(row[14]),
                    safe_float(row[15]),
                    safe_float(row[16]),
                    safe_float(row[17]),
                    safe_float(row[18]),
                    safe_float(row[19]),
                    normalize_date(row[20]),
                    row[21],
                    row[22],
                    row[23],
                    row[24],
                    str(input_path),
                    imported_at,
                ),
            )
            inserted += 1
        conn.commit()

    print(f"[batch] id={batch_id}")
    print(f"[output] db={db_path}")
    print(f"[result] inserted={inserted}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
