import argparse
import csv
import asyncio
import io
import json
import os
import random
import re
import sqlite3
import sys
import time
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
try:
    from playwright.async_api import async_playwright
except Exception:
    async_playwright = None


DEFAULT_COLUMNS = [
    "supplier_part_number",
    "normalized_part_number",
    "normalization_basis",
    "supplier_stock_raw",
    "supplier_stock_qty",
    "supplier_stock_year",
    "supplier_stock_lot",
    "supplier_package",
    "supplier_lead_time",
    "supplier_stock_note",
    "parse_confidence",
    "parse_status",
    "source_platform",
    "searched_keyword",
    "matched_part_number",
    "match_confidence",
    "price",
    "currency",
    "package",
    "moq",
    "stock",
    "seller_name",
    "region",
    "url",
    "capture_time",
    "raw_snapshot_path",
    "match_status",
    "notes",
]

PART_KEYWORDS = ["型号", "型號", "model", "part", "mpn", "料号", "料號", "pn"]
STOCK_KEYWORDS = [
    "库存",
    "庫存",
    "数量",
    "數量",
    "qty",
    "quantity",
    "batch",
    "lot",
    "lead",
    "交期",
    "备注",
    "備註",
    "note",
    "remark",
]

PRICE_PATTERNS = [
    re.compile(r"(?:USD|US\$|\$|CNY|RMB|￥|¥)\s*([0-9]+(?:\.[0-9]+)?)", re.IGNORECASE),
    re.compile(r"([0-9]+(?:\.[0-9]+)?)\s*(?:USD|US\$|CNY|RMB|￥|¥)", re.IGNORECASE),
]
MOQ_RE = re.compile(r"(?:MOQ|起订量|起訂量|最小起订量|最小起訂量)\s*[:：]?\s*([0-9,]+)", re.IGNORECASE)
STOCK_RE = re.compile(r"(?:库存|庫存|in\s*stock)\s*[:：]?\s*([0-9,]+)", re.IGNORECASE)
YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")
SHORT_YEAR_TOKEN_RE = re.compile(r"\b(?:20)?(\d{2})\+?\b")
LOT_RE = re.compile(r"\b(?:LOT|DC|批次|批號|批号)\s*[:：-]?\s*([A-Z0-9\-_/]+)", re.IGNORECASE)
LEAD_RE = re.compile(r"(?:交期|lead\s*time)\s*[:：-]?\s*([^|;,]+)", re.IGNORECASE)
PACKAGE_RE = re.compile(r"(?:封装|封裝|package|pkg)\s*[:：-]?\s*([A-Z0-9\-_/]+)", re.IGNORECASE)
PART_INLINE_RE = re.compile(r"^\s*([A-Za-z0-9][A-Za-z0-9\-_/]{3,})(?:\s+|[,;|/]+)?(.*)$")
BARE_PACKAGE_RE = re.compile(r"\b(SOP|SOIC|TSSOP|QFN|BGA|LQFP|QFP|DIP|TO-220|TO220|DFN|SOT-23|SOT23)\s*-?\s*([0-9]{1,3})?\b", re.IGNORECASE)


class PlatformQueryError(Exception):
    pass


class BommanCdpCookieClient:
    def __init__(self, cdp_endpoint: str, timeout_seconds: int):
        self.cdp_endpoint = cdp_endpoint
        self.timeout_ms = max(1000, timeout_seconds * 1000)
        self.loop = asyncio.new_event_loop()
        self.playwright = None
        self.browser = None
        self.context = None

    async def _start(self) -> None:
        if async_playwright is None:
            raise PlatformQueryError("playwright is not installed; cannot use --bomman-use-cdp-cookie")
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.connect_over_cdp(self.cdp_endpoint)
        if self.browser.contexts:
            self.context = self.browser.contexts[0]
        else:
            self.context = await self.browser.new_context()

    def start(self) -> None:
        self.loop.run_until_complete(self._start())

    async def _fetch_html(self, url: str) -> str:
        if self.context is None:
            raise PlatformQueryError("cdp context not ready")
        page = await self.context.new_page()
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=self.timeout_ms)
            await page.wait_for_timeout(1800)
            return await page.content()
        finally:
            await page.close()

    def fetch_html(self, url: str) -> str:
        return self.loop.run_until_complete(self._fetch_html(url))

    async def _close(self) -> None:
        try:
            if self.browser is not None:
                await self.browser.close()
        finally:
            if self.playwright is not None:
                await self.playwright.stop()

    def close(self) -> None:
        try:
            self.loop.run_until_complete(self._close())
        finally:
            self.loop.close()


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Vendor list platform price collector (bomman + so.szlcsc).")
    p.add_argument("--input", default=r"F:/Jay_ic_tw/data/vendor_parts_20260310.csv", help="Input vendor file path.")
    p.add_argument("--output-dir", default=r"F:/Jay_ic_tw/data", help="Output directory.")
    p.add_argument("--limit", type=int, default=0, help="Optional max vendor rows to process.")
    p.add_argument("--sleep-seconds", type=float, default=1.5, help="Delay between requests.")
    p.add_argument(
        "--sleep-between-model-min-seconds",
        type=float,
        default=30.0,
        help="Random wait min seconds after each model is completed.",
    )
    p.add_argument(
        "--sleep-between-model-max-seconds",
        type=float,
        default=50.0,
        help="Random wait max seconds after each model is completed.",
    )
    p.add_argument("--timeout-seconds", type=int, default=25, help="HTTP timeout in seconds.")
    p.add_argument("--resume-state-file", default=r"F:/Jay_ic_tw/data/vendor_parts_20260310_resume_state.json", help="Resume checkpoint JSON.")
    p.add_argument("--lock-file", default=r"F:/Jay_ic_tw/data/vendor_parts_20260310_task.lock", help="Single-instance lock file.")
    p.add_argument("--output-csv", default="", help="Optional explicit output csv path.")
    p.add_argument("--db-path", default=r"F:/Jay_ic_tw/sol.db", help="SQLite database path for structured persistence.")
    p.add_argument("--no-db-write", action="store_true", help="Disable SQLite writes for this run.")
    p.add_argument(
        "--bomman-use-cdp-cookie",
        action="store_true",
        help="Use local Chrome via CDP for bomman.com so logged-in cookies are reused.",
    )
    p.add_argument(
        "--cdp-endpoint",
        default="http://127.0.0.1:9222",
        help="CDP endpoint for existing local Chrome.",
    )
    p.add_argument(
        "--disable-env-proxy",
        action="store_true",
        help="Disable HTTP_PROXY/HTTPS_PROXY only for this run (use local network IP).",
    )
    return p.parse_args()


def _is_pid_alive(pid: int) -> bool:
    if pid <= 0:
        return False
    try:
        os.kill(pid, 0)
        return True
    except OSError:
        return False


def acquire_lock(lock_path: Path) -> None:
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    pid = os.getpid()
    try:
        fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(str(pid))
        return
    except FileExistsError:
        old_pid = 0
        try:
            txt = lock_path.read_text(encoding="utf-8").strip()
            old_pid = int(txt) if txt.isdigit() else 0
        except Exception:
            old_pid = 0
        if old_pid and _is_pid_alive(old_pid):
            raise RuntimeError(f"another task is running (pid={old_pid}), lock={lock_path}")
        lock_path.unlink(missing_ok=True)
        fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(str(pid))


def release_lock(lock_path: Path) -> None:
    try:
        if lock_path.exists():
            lock_path.unlink(missing_ok=True)
    except Exception:
        pass


def normalize_part_number(raw: str) -> str:
    text = (raw or "").strip().upper()
    text = re.sub(r"\s+", "", text)
    return text


def normalization_basis(raw: str, normalized: str) -> str:
    stripped = (raw or "").strip()
    if not stripped:
        return "empty"
    if stripped == normalized:
        return "exact_input"
    if stripped.upper() == normalized:
        return "normalized_by_case"
    compact = re.sub(r"\s+", "", stripped.upper())
    if compact == normalized:
        return "normalized_by_rule"
    return "manual_review"


def normalize_token_text(raw: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", (raw or "").upper())


def safe_float(text: str) -> float | None:
    if text is None:
        return None
    s = str(text).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def read_vendor_rows(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    raw = path.read_bytes()
    rows: list[dict[str, Any]] = []

    # Case A: extension says csv, but payload is xlsx (PK magic).
    if raw[:2] == b"PK":
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return [], {"format": "xlsx-in-csv-name", "sheet": wb.sheetnames[0], "headers": []}
        headers = ["" if x is None else str(x).strip() for x in values[0]]
        for row in values[1:]:
            rec = {}
            for i, h in enumerate(headers):
                key = h if h else f"col_{i+1}"
                rec[key] = "" if i >= len(row) or row[i] is None else str(row[i]).strip()
            rows.append(rec)
        return rows, {"format": "xlsx-in-csv-name", "sheet": wb.sheetnames[0], "headers": headers}

    # Case B: real csv text.
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    for row in reader:
        rows.append({(k or "").strip(): (v or "").strip() for k, v in row.items()})
    return rows, {"format": "csv", "headers": headers}


def choose_column(headers: list[str], keywords: list[str]) -> str | None:
    scored: list[tuple[int, str]] = []
    for h in headers:
        hl = (h or "").lower()
        score = 0
        for kw in keywords:
            if kw.lower() in hl:
                score += len(kw)
        if score > 0:
            scored.append((score, h))
    if not scored:
        return None
    scored.sort(reverse=True)
    return scored[0][1]


def detect_vendor_fields(rows: list[dict[str, Any]]) -> dict[str, Any]:
    headers = list(rows[0].keys()) if rows else []
    part_col = choose_column(headers, PART_KEYWORDS)
    stock_cols = [h for h in headers if any(kw.lower() in (h or "").lower() for kw in STOCK_KEYWORDS)]
    inline_stock = False
    if not part_col and headers:
        part_col = headers[0]
    if not stock_cols and len(headers) > 1:
        stock_cols = headers[1:]
    if part_col and not stock_cols:
        samples = [str(r.get(part_col, "")).strip() for r in rows[:10]]
        inline_hits = 0
        for sample in samples:
            m = PART_INLINE_RE.match(sample)
            if m and m.group(2).strip():
                inline_hits += 1
        inline_stock = inline_hits > 0
    return {
        "part_col": part_col,
        "stock_cols": stock_cols,
        "headers": headers,
        "inline_stock": inline_stock,
    }


def split_inline_part_and_stock(raw_value: str) -> tuple[str, str]:
    text = (raw_value or "").strip()
    m = PART_INLINE_RE.match(text)
    if not m:
        return text, ""
    return m.group(1).strip(), m.group(2).strip()


def infer_qty(text: str) -> float | None:
    qty_patterns = [
        re.compile(r"(?:库存|庫存|数量|數量|qty|quantity)\s*[:：-]?\s*([0-9][0-9,]*(?:\.[0-9]+)?)\s*([kKmM]{0,2})", re.IGNORECASE),
        re.compile(r"\b([0-9][0-9,]*(?:\.[0-9]+)?)\s*(pcs|pc|k|kk|m)?\b", re.IGNORECASE),
    ]
    for rx in qty_patterns:
        m = rx.search(text)
        if m:
            qty = safe_float(m.group(1))
            if qty is None:
                continue
            unit = (m.group(2) or "").lower() if len(m.groups()) >= 2 else ""
            if unit == "k":
                qty *= 1000
            elif unit == "kk":
                qty *= 1000000
            elif unit == "m":
                qty *= 1000000
            return qty
    return None


def infer_year(raw_text: str) -> str:
    year_m = YEAR_RE.search(raw_text)
    if year_m:
        return year_m.group(0)
    raw_upper = raw_text.upper()
    dc_m = re.search(r"\bDC\s*[:：-]?\s*(\d{2,4}\+?)", raw_upper)
    if dc_m:
        token = dc_m.group(1).rstrip("+")
        if len(token) == 2:
            return f"20{token}"
        if len(token) == 4:
            return token
    return ""


def infer_package(raw_text: str) -> str:
    package_m = PACKAGE_RE.search(raw_text)
    if package_m:
        return package_m.group(1).strip()
    bare_m = BARE_PACKAGE_RE.search(raw_text)
    if bare_m:
        prefix = bare_m.group(1).upper().replace("TO220", "TO-220").replace("SOT23", "SOT-23")
        suffix = bare_m.group(2) or ""
        return f"{prefix}{suffix}"
    return ""


def parse_supplier_stock_fields(raw_text: str, stock_data: dict[str, str]) -> dict[str, Any]:
    year = infer_year(raw_text)

    lot = ""
    lot_m = LOT_RE.search(raw_text)
    if lot_m:
        lot = lot_m.group(1).strip()

    lead = ""
    lead_m = LEAD_RE.search(raw_text)
    if lead_m:
        lead = lead_m.group(1).strip()

    package = infer_package(raw_text)

    qty = infer_qty(raw_text)
    if qty is None:
        for key, value in stock_data.items():
            if any(x in key.lower() for x in ["库存", "庫存", "数量", "數量", "qty", "quantity"]):
                qty = safe_float(value)
                if qty is not None:
                    break

    note_segments: list[str] = []
    for key, value in stock_data.items():
        vv = (value or "").strip()
        if vv:
            note_segments.append(f"{key}:{vv}")
    parse_hits = sum(1 for value in (qty, year, lot, package, lead) if value)
    parse_conf = min(1.0, 0.25 + parse_hits * 0.15) if raw_text else 1.0
    parse_status = "parsed" if parse_hits > 0 else ("raw_only" if raw_text else "empty")
    return {
        "supplier_stock_qty": "" if qty is None else str(int(qty) if qty.is_integer() else qty),
        "supplier_stock_year": year,
        "supplier_stock_lot": lot,
        "supplier_package": package,
        "supplier_lead_time": lead,
        "supplier_stock_note": " | ".join(note_segments),
        "parse_confidence": f"{parse_conf:.2f}",
        "parse_status": parse_status,
    }


def extract_supplier_stock(
    stock_data: dict[str, str],
    inline_raw: str = "",
) -> dict[str, Any]:
    raw_pairs = [f"{k}={v}" for k, v in stock_data.items() if str(v).strip()]
    if inline_raw:
        raw_pairs.insert(0, f"inline={inline_raw}")
    raw_text = " | ".join(raw_pairs)
    parsed = parse_supplier_stock_fields(raw_text, stock_data)
    parsed["supplier_stock_raw"] = raw_text
    return parsed


def fetch_html(url: str, timeout_seconds: int, user_agent: str) -> str:
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": user_agent,
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        },
    )
    with urllib.request.urlopen(req, timeout=timeout_seconds) as resp:
        raw = resp.read()
        content_type = (resp.headers.get("Content-Type") or "").lower()
    charset = "utf-8"
    m = re.search(r"charset=([a-zA-Z0-9_-]+)", content_type)
    if m:
        charset = m.group(1)
    try:
        return raw.decode(charset, errors="replace")
    except LookupError:
        return raw.decode("utf-8", errors="replace")


def pick_currency(text: str) -> str:
    t = text.upper()
    if "USD" in t or "US$" in t or "$" in t:
        return "USD"
    if "CNY" in t or "RMB" in t or "￥" in text or "¥" in text:
        return "CNY"
    return ""


def candidate_lines_from_html(html: str) -> list[str]:
    cleaned = re.sub(r"<script[\s\S]*?</script>", " ", html, flags=re.IGNORECASE)
    cleaned = re.sub(r"<style[\s\S]*?</style>", " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"<[^>]+>", "\n", cleaned)
    lines = [re.sub(r"\s+", " ", ln).strip() for ln in cleaned.splitlines()]
    return [ln for ln in lines if ln]


def save_snapshot(output_dir: Path, platform: str, part: str, html: str, lines: list[str]) -> str:
    snapshot_dir = output_dir / "platform_snapshots"
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_platform = re.sub(r"[^A-Za-z0-9]+", "_", platform).strip("_")
    safe_part = re.sub(r"[^A-Za-z0-9]+", "_", part).strip("_")
    path = snapshot_dir / f"{stamp}_{safe_platform}_{safe_part}.txt"
    snippet = "\n".join(lines[:120])
    payload = f"platform={platform}\npart={part}\ncapture_time={datetime.now().isoformat(timespec='seconds')}\n\n{snippet}\n"
    path.write_text(payload, encoding="utf-8")
    return str(path)


def determine_match_confidence(part: str, hit_lines: list[str], matched_part: str) -> float:
    if not hit_lines:
        return 0.0
    if matched_part and normalize_token_text(matched_part) == normalize_token_text(part) and len(hit_lines) == 1:
        return 1.0
    if matched_part and len(hit_lines) == 1:
        return 0.9
    if len(hit_lines) > 1:
        return 0.45
    return 0.2


def extract_matched_part(primary: str, part: str) -> str:
    candidates = re.findall(r"\b[A-Za-z0-9][A-Za-z0-9\-_/]{3,}\b", primary)
    norm_part = normalize_token_text(part)
    for candidate in candidates:
        if normalize_token_text(candidate) == norm_part:
            return candidate.upper()
    return ""


def parse_platform_result(part: str, platform: str, url: str, html: str, output_dir: Path) -> dict[str, Any]:
    lines = candidate_lines_from_html(html)
    part_u = normalize_token_text(part)
    snapshot_path = save_snapshot(output_dir, platform, part, html, lines)
    capture_time = datetime.now().isoformat(timespec="seconds")

    hit_lines = [ln for ln in lines if part_u and part_u in normalize_token_text(ln)]

    if not hit_lines:
        return {
            "source_platform": platform,
            "searched_keyword": part,
            "matched_part_number": "",
            "match_confidence": "0.00",
            "price": "",
            "currency": "",
            "package": "",
            "moq": "",
            "stock": "",
            "seller_name": "",
            "region": "",
            "url": url,
            "capture_time": capture_time,
            "raw_snapshot_path": snapshot_path,
            "match_status": "not_found",
            "notes": "no part token hit in response text",
        }

    match_status = "matched_exact" if len(hit_lines) == 1 else "multiple_candidates"
    primary = hit_lines[0]
    price_m = None
    for rx in PRICE_PATTERNS:
        price_m = rx.search(primary)
        if price_m:
            break
    moq_m = MOQ_RE.search(primary)
    stock_m = STOCK_RE.search(primary)

    matched_part = extract_matched_part(primary, part)
    price = price_m.group(1) if price_m else ""
    currency = pick_currency(primary)
    moq = moq_m.group(1).replace(",", "") if moq_m else ""
    stock = stock_m.group(1).replace(",", "") if stock_m else ""
    confidence = determine_match_confidence(part, hit_lines, matched_part)

    notes = []
    if match_status == "multiple_candidates":
        notes.append(f"candidate_count={len(hit_lines)}")
    if not price:
        notes.append("price_not_parsed")
    if re.search(r"(captcha|forbidden|access denied|blocked)", html, flags=re.IGNORECASE):
        match_status = "blocked"
        confidence = 0.0
        notes.append("platform_block_signal")

    return {
        "source_platform": platform,
        "searched_keyword": part,
        "matched_part_number": matched_part,
        "match_confidence": f"{confidence:.2f}",
        "price": price,
        "currency": currency,
        "package": "",
        "moq": moq,
        "stock": stock,
        "seller_name": "",
        "region": "",
        "url": url,
        "capture_time": capture_time,
        "raw_snapshot_path": snapshot_path,
        "match_status": match_status if price and match_status != "blocked" else ("blocked" if match_status == "blocked" else "manual_review"),
        "notes": " | ".join(notes),
    }


def query_bomman(
    part: str,
    timeout_seconds: int,
    user_agent: str,
    output_dir: Path,
    bomman_cdp_client: BommanCdpCookieClient | None = None,
) -> dict[str, Any]:
    q = urllib.parse.quote(part)
    # Known searchable endpoint pattern (fallback-able via query URL itself).
    url = f"https://www.bomman.com/?s={q}"
    if bomman_cdp_client is not None:
        html = bomman_cdp_client.fetch_html(url)
    else:
        html = fetch_html(url, timeout_seconds, user_agent)
    return parse_platform_result(part, "bomman.com", url, html, output_dir)


def query_szlcsc(part: str, timeout_seconds: int, user_agent: str, output_dir: Path) -> dict[str, Any]:
    q = urllib.parse.quote(part)
    url = f"https://so.szlcsc.com/global.html?k={q}"
    html = fetch_html(url, timeout_seconds, user_agent)
    return parse_platform_result(part, "so.szlcsc.com", url, html, output_dir)


def load_resume(state_path: Path) -> dict[str, Any]:
    if not state_path.exists():
        return {}
    try:
        return json.loads(state_path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_resume(state_path: Path, payload: dict[str, Any]) -> None:
    state_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def ensure_output_csv(path: Path) -> None:
    if path.exists() and path.stat().st_size > 0:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=DEFAULT_COLUMNS)
        w.writeheader()


def append_rows(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("a", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=DEFAULT_COLUMNS)
        for row in rows:
            payload = {k: row.get(k, "") for k in DEFAULT_COLUMNS}
            w.writerow(payload)


def load_completed_models_from_output(path: Path) -> set[str]:
    if not path.exists() or path.stat().st_size == 0:
        return set()
    counts: dict[str, int] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            model = normalize_part_number(row.get("normalized_part_number", ""))
            if not model:
                model = normalize_part_number(row.get("supplier_part_number", ""))
            if not model:
                continue
            counts[model] = counts.get(model, 0) + 1
    # A model is considered completed when both platform rows already exist.
    return {m for m, c in counts.items() if c >= 2}


def connect_db(db_path: Path) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_phase1_tables(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS supplier_items (
          supplier_item_id INTEGER PRIMARY KEY AUTOINCREMENT,
          batch_id TEXT NOT NULL,
          supplier_name TEXT,
          supplier_part_number TEXT NOT NULL,
          normalized_part_number TEXT,
          normalization_basis TEXT,
          supplier_stock_raw TEXT,
          supplier_stock_qty REAL,
          supplier_stock_year TEXT,
          supplier_stock_lot TEXT,
          supplier_package TEXT,
          supplier_lead_time TEXT,
          supplier_stock_note TEXT,
          parse_confidence REAL,
          parse_status TEXT,
          created_at TEXT NOT NULL,
          UNIQUE(batch_id, supplier_part_number, supplier_stock_raw)
        );

        CREATE TABLE IF NOT EXISTS market_quotes (
          market_quote_id INTEGER PRIMARY KEY AUTOINCREMENT,
          batch_id TEXT NOT NULL,
          supplier_item_id INTEGER NOT NULL,
          source_platform TEXT NOT NULL,
          searched_keyword TEXT,
          matched_part_number TEXT,
          match_confidence REAL,
          price REAL,
          currency TEXT,
          package TEXT,
          moq REAL,
          stock REAL,
          seller_name TEXT,
          region TEXT,
          url TEXT,
          capture_time TEXT,
          match_status TEXT,
          notes TEXT,
          raw_snapshot_path TEXT,
          created_at TEXT NOT NULL,
          UNIQUE(batch_id, supplier_item_id, source_platform),
          FOREIGN KEY(supplier_item_id) REFERENCES supplier_items(supplier_item_id)
        );
        """
    )
    conn.commit()


def nullable_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return safe_float(text)


def upsert_supplier_item(conn: sqlite3.Connection, batch_id: str, row: dict[str, Any]) -> int:
    created_at = datetime.now().isoformat(timespec="seconds")
    conn.execute(
        """
        INSERT INTO supplier_items(
          batch_id, supplier_name, supplier_part_number, normalized_part_number, normalization_basis,
          supplier_stock_raw, supplier_stock_qty, supplier_stock_year, supplier_stock_lot,
          supplier_package, supplier_lead_time, supplier_stock_note, parse_confidence,
          parse_status, created_at
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(batch_id, supplier_part_number, supplier_stock_raw) DO UPDATE SET
          normalized_part_number=excluded.normalized_part_number,
          normalization_basis=excluded.normalization_basis,
          supplier_stock_qty=excluded.supplier_stock_qty,
          supplier_stock_year=excluded.supplier_stock_year,
          supplier_stock_lot=excluded.supplier_stock_lot,
          supplier_package=excluded.supplier_package,
          supplier_lead_time=excluded.supplier_lead_time,
          supplier_stock_note=excluded.supplier_stock_note,
          parse_confidence=excluded.parse_confidence,
          parse_status=excluded.parse_status
        """,
        (
            batch_id,
            "",
            row.get("supplier_part_number", ""),
            row.get("normalized_part_number", ""),
            row.get("normalization_basis", ""),
            row.get("supplier_stock_raw", ""),
            nullable_float(row.get("supplier_stock_qty", "")),
            row.get("supplier_stock_year", ""),
            row.get("supplier_stock_lot", ""),
            row.get("supplier_package", ""),
            row.get("supplier_lead_time", ""),
            row.get("supplier_stock_note", ""),
            nullable_float(row.get("parse_confidence", "")),
            row.get("parse_status", ""),
            created_at,
        ),
    )
    item_id = conn.execute(
        """
        SELECT supplier_item_id
        FROM supplier_items
        WHERE batch_id=? AND supplier_part_number=? AND supplier_stock_raw=?
        """,
        (batch_id, row.get("supplier_part_number", ""), row.get("supplier_stock_raw", "")),
    ).fetchone()
    conn.commit()
    return int(item_id["supplier_item_id"])


def upsert_market_quote(conn: sqlite3.Connection, batch_id: str, supplier_item_id: int, row: dict[str, Any]) -> None:
    created_at = datetime.now().isoformat(timespec="seconds")
    conn.execute(
        """
        INSERT INTO market_quotes(
          batch_id, supplier_item_id, source_platform, searched_keyword, matched_part_number,
          match_confidence, price, currency, package, moq, stock, seller_name, region,
          url, capture_time, match_status, notes, raw_snapshot_path, created_at
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(batch_id, supplier_item_id, source_platform) DO UPDATE SET
          searched_keyword=excluded.searched_keyword,
          matched_part_number=excluded.matched_part_number,
          match_confidence=excluded.match_confidence,
          price=excluded.price,
          currency=excluded.currency,
          package=excluded.package,
          moq=excluded.moq,
          stock=excluded.stock,
          seller_name=excluded.seller_name,
          region=excluded.region,
          url=excluded.url,
          capture_time=excluded.capture_time,
          match_status=excluded.match_status,
          notes=excluded.notes,
          raw_snapshot_path=excluded.raw_snapshot_path
        """,
        (
            batch_id,
            supplier_item_id,
            row.get("source_platform", ""),
            row.get("searched_keyword", ""),
            row.get("matched_part_number", ""),
            nullable_float(row.get("match_confidence", "")),
            nullable_float(row.get("price", "")),
            row.get("currency", ""),
            row.get("package", ""),
            nullable_float(row.get("moq", "")),
            nullable_float(row.get("stock", "")),
            row.get("seller_name", ""),
            row.get("region", ""),
            row.get("url", ""),
            row.get("capture_time", ""),
            row.get("match_status", ""),
            row.get("notes", ""),
            row.get("raw_snapshot_path", ""),
            created_at,
        ),
    )
    conn.commit()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    output_dir = Path(args.output_dir)
    state_path = Path(args.resume_state_file)
    lock_path = Path(args.lock_file)
    db_path = Path(args.db_path)

    if args.disable_env_proxy:
        for k in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"):
            os.environ.pop(k, None)
        print("[network] env proxy disabled for this run")

    if not input_path.exists():
        print(f"[error] input not found: {input_path}")
        return 2

    acquire_lock(lock_path)
    bomman_cdp_client: BommanCdpCookieClient | None = None
    db_conn: sqlite3.Connection | None = None
    try:
        rows, input_meta = read_vendor_rows(input_path)
        if not rows:
            print("[error] input has no data rows")
            return 3

        detected = detect_vendor_fields(rows)
        part_col = detected["part_col"]
        stock_cols = detected["stock_cols"]
        inline_stock = bool(detected.get("inline_stock"))
        if not part_col:
            print(f"[error] could not detect part number column. headers={detected['headers']}")
            return 4

        resume = load_resume(state_path)
        start_index = int(resume.get("next_index", 0))
        if start_index < 0 or start_index > len(rows):
            start_index = 0
        batch_id = str(resume.get("batch_id", "")).strip()
        if not batch_id:
            batch_id = datetime.now().strftime("vendor_batch_%Y%m%d_%H%M%S")

        if args.output_csv:
            out_csv = Path(args.output_csv)
        else:
            out_csv_str = resume.get("output_csv", "")
            if out_csv_str:
                out_csv = Path(out_csv_str)
            else:
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                out_csv = output_dir / f"vendor_platform_prices_{stamp}.csv"

        ensure_output_csv(out_csv)
        completed_models = load_completed_models_from_output(out_csv)
        if not args.no_db_write:
            db_conn = connect_db(db_path)
            ensure_phase1_tables(db_conn)

        print(f"[input] file={input_path}")
        print(f"[input] format={input_meta.get('format')} headers={input_meta.get('headers')}")
        print(f"[detect] part_col={part_col} stock_cols={stock_cols} inline_stock={inline_stock}")
        print(f"[resume] next_index={start_index} total_rows={len(rows)}")
        print(f"[batch] id={batch_id}")
        print(f"[output] csv={out_csv}")
        print(f"[output] db={db_path} enabled={not args.no_db_write}")
        print(f"[skip] completed_models_in_output={len(completed_models)}")

        user_agent = (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/123.0.0.0 Safari/537.36"
        )
        if args.bomman_use_cdp_cookie:
            bomman_cdp_client = BommanCdpCookieClient(args.cdp_endpoint, args.timeout_seconds)
            try:
                bomman_cdp_client.start()
            except Exception as exc:
                print(f"[error] failed to connect local chrome cookie session: {exc}")
                return 5
            print(f"[bomman] using local chrome cookie via cdp={args.cdp_endpoint}")

        end_index = len(rows)
        if args.limit and args.limit > 0:
            end_index = min(end_index, start_index + args.limit)

        wait_min = args.sleep_between_model_min_seconds
        wait_max = args.sleep_between_model_max_seconds
        if wait_min < 0:
            wait_min = 0.0
        if wait_max < wait_min:
            wait_max = wait_min

        for idx in range(start_index, end_index):
            src = rows[idx]
            supplier_part_raw = str(src.get(part_col, "")).strip()
            inline_raw = ""
            supplier_part = supplier_part_raw
            if inline_stock:
                supplier_part, inline_raw = split_inline_part_and_stock(supplier_part_raw)
            normalized = normalize_part_number(supplier_part)
            if normalized in completed_models:
                next_idx = idx + 1
                save_resume(
                    state_path,
                    {
                        "input_file": str(input_path),
                        "input_format": input_meta.get("format"),
                        "part_col": part_col,
                        "stock_cols": stock_cols,
                        "output_csv": str(out_csv),
                        "db_path": str(db_path),
                        "batch_id": batch_id,
                        "next_index": next_idx,
                        "updated_at": datetime.now().isoformat(timespec="seconds"),
                    },
                )
                print(f"[{idx+1}/{end_index}] part={normalized} skipped(already_completed)")
                continue

            stock_data = {c: str(src.get(c, "")).strip() for c in stock_cols}
            parsed_stock = extract_supplier_stock(stock_data, inline_raw=inline_raw)

            base = {
                "supplier_part_number": supplier_part,
                "normalized_part_number": normalized,
                "normalization_basis": normalization_basis(supplier_part, normalized),
            }
            base.update(parsed_stock)

            platform_rows: list[dict[str, Any]] = []
            for platform_fn in (query_bomman, query_szlcsc):
                try:
                    if platform_fn is query_bomman:
                        r = platform_fn(
                            normalized,
                            args.timeout_seconds,
                            user_agent,
                            output_dir,
                            bomman_cdp_client=bomman_cdp_client,
                        )
                    else:
                        r = platform_fn(normalized, args.timeout_seconds, user_agent, output_dir)
                except Exception as exc:
                    platform_name = "bomman.com" if platform_fn is query_bomman else "so.szlcsc.com"
                    r = {
                        "source_platform": platform_name,
                        "searched_keyword": normalized,
                        "matched_part_number": "",
                        "match_confidence": "0.00",
                        "price": "",
                        "currency": "",
                        "package": "",
                        "moq": "",
                        "stock": "",
                        "seller_name": "",
                        "region": "",
                        "url": "",
                        "capture_time": datetime.now().isoformat(timespec="seconds"),
                        "raw_snapshot_path": "",
                        "match_status": "fetch_error",
                        "notes": str(exc).splitlines()[0][:240],
                    }
                row = {**base, **r}
                platform_rows.append(row)
                time.sleep(max(0.0, args.sleep_seconds))

            append_rows(out_csv, platform_rows)
            if db_conn is not None:
                supplier_item_id = upsert_supplier_item(db_conn, batch_id, base)
                for row in platform_rows:
                    upsert_market_quote(db_conn, batch_id, supplier_item_id, row)
            completed_models.add(normalized)

            next_idx = idx + 1
            save_resume(
                state_path,
                {
                    "input_file": str(input_path),
                    "input_format": input_meta.get("format"),
                    "part_col": part_col,
                        "stock_cols": stock_cols,
                        "output_csv": str(out_csv),
                        "db_path": str(db_path),
                        "batch_id": batch_id,
                        "next_index": next_idx,
                        "updated_at": datetime.now().isoformat(timespec="seconds"),
                    },
            )
            print(f"[{idx+1}/{end_index}] part={normalized} done")
            if idx + 1 < end_index:
                pause = random.uniform(wait_min, wait_max)
                print(f"[wait] sleeping {pause:.1f}s before next model")
                time.sleep(pause)

        print("[done] completed run")
        return 0
    finally:
        if bomman_cdp_client is not None:
            bomman_cdp_client.close()
        if db_conn is not None:
            db_conn.close()
        release_lock(lock_path)


if __name__ == "__main__":
    sys.exit(main())
